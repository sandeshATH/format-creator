from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from excel_form import extract_fields, fill_template


def build_workbook(labels, values):
    wb = Workbook()
    ws = wb.active
    for row, label in enumerate(labels, start=1):
        label_cell = ws.cell(row=row, column=1, value=label)
        label_cell.font = Font(bold=True)
        value_cell = ws.cell(row=row, column=5, value=values[row - 1])
        value_cell.font = Font(bold=True)
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def test_extract_fields_returns_bold_column_e_values_with_labels():
    labels = ["Name", "Email"]
    values = ["Ada", "ada@example.com"]
    workbook_bytes = build_workbook(labels, values)

    result = extract_fields(BytesIO(workbook_bytes))

    assert [field["label"] for field in result] == labels
    assert [field["value"] for field in result] == values


def test_extract_fields_raises_when_no_bold_fields():
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="Name")
    buffer = BytesIO()
    wb.save(buffer)
    workbook_bytes = buffer.getvalue()

    with pytest.raises(ValueError):
        extract_fields(BytesIO(workbook_bytes))


def test_fill_template_writes_data_into_field_coordinates():
    labels = ["Name", "Email"]
    values = ["Ada", "ada@example.com"]
    workbook_bytes = build_workbook(labels, values)
    fields = extract_fields(BytesIO(workbook_bytes))

    filled_bytes = fill_template(workbook_bytes, fields, ["Grace", "grace@example.com"])

    workbook = load_workbook(BytesIO(filled_bytes))
    sheet = workbook.active

    assert sheet.cell(row=1, column=5).value == "Grace"
    assert sheet.cell(row=2, column=5).value == "grace@example.com"
