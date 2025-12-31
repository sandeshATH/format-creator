from __future__ import annotations

from io import BytesIO
from typing import Iterable, List, Dict

from openpyxl import load_workbook


def extract_fields(stream: BytesIO) -> List[Dict[str, object]]:
    """Read editable bold cells and return field metadata.

    Raises:
        ValueError: If no editable fields are found.
    """

    workbook = load_workbook(stream, data_only=True)
    sheet = workbook.active

    fields: List[Dict[str, object]] = []
    for row in range(1, sheet.max_row + 1):
        label_cell = sheet.cell(row=row, column=1)
        value_cell = sheet.cell(row=row, column=5)
        if label_cell.value is None:
            continue
        label = str(label_cell.value).strip()
        if not label:
            continue
        if not value_cell.font or not value_cell.font.bold:
            continue
        fields.append(
            {
                "key": f"r{row}c{value_cell.column}",
                "label": label,
                "row": row,
                "column": value_cell.column,
                "value": "" if value_cell.value is None else str(value_cell.value),
            }
        )

    if not fields:
        raise ValueError("No editable bold fields were found in the template.")

    return fields


def fill_template(
    template_bytes: bytes, fields: Iterable[Dict[str, object]], values: Iterable[str]
) -> bytes:
    """Fill the first worksheet of a template workbook with provided field values.

    The input workbook is not modified; a new workbook is returned as bytes.
    """

    workbook = load_workbook(BytesIO(template_bytes))
    sheet = workbook.active

    for field, value in zip(fields, values):
        row = int(field["row"])
        column = int(field["column"])
        sheet.cell(row=row, column=column, value=value)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
